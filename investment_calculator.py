#! usr/bin/env python3
import sys
import openpyxl
from datetime import datetime
from constants import holding_locations, CAD_CASH, CAD_TFSA, format_sheet

#USAGE checker
if len(sys.argv) != 3:
    print('USAGE: Input 1: amount wanting to invest. Input 2: Amount into TFSA.')
    exit()

#Constants for portfolio proportions.
CDN_BOND_PCNT = 20
CDN_INDX_PCNT = 26
USA_INDX_PCNT = 27
INTL_INDX_PCNT = 27

TOTAL = 'F'

accounts = [CAD_CASH, CAD_TFSA]

NEW_CASH = '6'
NEW_TFSA = '7'

wb = openpyxl.load_workbook('investments.xlsx')

# Read excel doc. Need to scrape values for each account and each holding. (8 total)
names = wb.sheetnames

#sheet represents the last investment.
sheet = wb[names[0]] #Open the 1st sheet since the newest is always first

account_amounts = []

for account in accounts:
    curr = []
    for holding in holding_locations:
        cell = sheet[holding + account]
        curr.append(cell.value)
    account_amounts.append(curr)
print(account_amounts)

current_totals = []
for i in range (len(holding_locations)):
    current_totals.append(account_amounts[0][i] + account_amounts[1][i])
       
#now account_amounts represents the ammount for each holding in each account

# Calculate amount going into each account
total_invested = sum(current_totals)
print(total_invested)

new_total = int(sys.argv[1]) + total_invested #add new amount to amount already in account

new_amnt_BOND = new_total * CDN_BOND_PCNT //100
new_amnt_CAD_INDX = new_total * CDN_INDX_PCNT //100
new_amnt_USA_INDX = new_total * USA_INDX_PCNT //100
new_amnt_INTL_INDX = new_total * INTL_INDX_PCNT //100
after_investing_amounts = [new_amnt_BOND, new_amnt_CAD_INDX, new_amnt_USA_INDX, new_amnt_INTL_INDX]
print(after_investing_amounts)

#now have the values that we want to have after investing. 
#Need to calculate the difference so that we can see how much to invest.
cash_changes = [0, 0, 0, 0]
tfsa_changes = [0, 0, 0, 0]

if (sys.argv[2] == '0'):
    #Not investing into TFSA, only concerned with cash account.
    for i in range(len(holding_locations)): #iterate over every holding
        cash_changes[i] = round(after_investing_amounts[i] - account_amounts[0][i] - account_amounts[1][i], 2)
else:
    #Only adding amount to the TFSA
    for i in range(len(holding_locations)): #iterate over every holding
        tfsa_changes[i] = round(-account_amounts[0][i] - account_amounts[1][i] + after_investing_amounts[i], 2)

print(cash_changes)
print (tfsa_changes)

# Write new values into new sheet.
now = datetime.now()
new_sheet_name = now.strftime("%d_%m_%Y")
new_sheet = wb.create_sheet(index = 0, title = new_sheet_name) #creates a new sheet with the current day.

format_sheet(new_sheet, new_sheet_name)

for i in range(len(cash_changes)):
    account_amounts[0][i] += cash_changes[i]

for i in range(len(tfsa_changes)):
    account_amounts[1][i] += tfsa_changes[i]

#Next we actually write the new amounts to the sheet

for i in range(len(accounts)):
    account = accounts[i]
    for j in range(len(holding_locations)):
        holding = holding_locations[j]
        new_sheet[holding + account] = account_amounts[i][j]

# write the changes to the sheet.

for i in range(len(holding_locations)):
    new_sheet[holding_locations[i] + NEW_CASH] = cash_changes[i]
    new_sheet[holding_locations[i] + NEW_TFSA] = tfsa_changes[i]

#write add the total changes added
new_sheet[TOTAL + NEW_CASH] = sum(cash_changes)
new_sheet[TOTAL + NEW_TFSA] = sum(tfsa_changes)

new_sheet[TOTAL + CAD_CASH] = sum(account_amounts[0])
new_sheet[TOTAL + CAD_TFSA] = sum(account_amounts[1])

#save the sheet.
wb.save('investments.xlsx')


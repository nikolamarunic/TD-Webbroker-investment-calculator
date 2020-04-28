#! usr/bin/env python3
import sys
import os
import openpyxl
from datetime import datetime
from constants import format_sheet, holding_locations, CAD_CASH, CAD_TFSA

#constants for rows where holdings are saved in spreadsheet
CDN_INDX = '9'
US_INDX = '10'
CDN_BOND = '11'
INTL_INDX = '12'
holding_locations_td_file = [CDN_INDX, US_INDX, CDN_BOND, INTL_INDX]

MARKET_VALUE ='H'

def get_data(sheet):
    values = [0, 0, 0, 0]
    for i in range(len(holding_locations_td_file)):
        holding = holding_locations_td_file[i]
        cell = sheet[MARKET_VALUE + holding]
        if cell.value: #nonempty cell
            values[i] = cell.value
        else:
            values[i] = 0
    return values

if not os.path.exists('CAD_CASH.xlsx') or not os.path.exists('CAD_TFSA.xlsx'):
    print('Td account data not found. Using previous investment data instead.')
    exit()

#Scrape data from the cash account
wb = openpyxl.load_workbook('CAD_CASH.xlsx')
sheet = wb[wb.sheetnames[0]]
cash_account_values = get_data(sheet)


#Scrape data from the TFSA
wb = openpyxl.load_workbook('CAD_TFSA.xlsx')
sheet = wb[wb.sheetnames[0]]
tfsa_account_values = get_data(sheet)


# Write data to the investments excel file
wb = openpyxl.load_workbook('investments.xlsx')

now = datetime.now()
new_sheet_name = "Holdings on " + now.strftime("%d_%m_%Y")
new_sheet = wb.create_sheet(index = 0, title = new_sheet_name) #creates a new sheet with the current day.

format_sheet(new_sheet, new_sheet_name)

for i in range(len(holding_locations)): #defined in the investment_calculator
    new_sheet[holding_locations[i] + CAD_CASH] = cash_account_values[i]
    new_sheet[holding_locations[i] + CAD_TFSA] = tfsa_account_values[i]

wb.save('investments.xlsx')

